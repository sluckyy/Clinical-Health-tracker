"""Parse the real July/August 2026 GP roster (old model) into
actuals.json - per-day assignments plus per-doctor statistics - in the same
shape roster_solver.py emits, so the dashboard can render actual and
generated months side by side.

Source layout (Sheet1): July grid in columns C..AG (days 1-31), August in
AI..BM. Rows: CB ED / CB GPA / CB OBS-oncall(GPO) / CB obs-support,
JT ED / JT GPA / JT OBS-oncall(GPO) / JT obs-support, Laura ED, Orroroo ED.
Weekday labels come from row 2 (the sheet's own labels).
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent

# canonical names: fixes typos and strips the "(SJ)"/"(HD)" backup notes
CANON = {
    "rick chalwin": "Rick Chalwin", "holly deer": "Holly Deer",
    "josh smith": "Josh Smith", "nitya sukheja": "Nitya Sukheja",
    "gareth o'reilly": "Gareth O'Reilly", "kc chen": "KC Chen",
    "rowena conway": "Rowena Conway", "simon jackson": "Simon Jackson",
    "niall fostyk": "Niall Fostyk", "krishna bayley": "Krishna Bayley",
    "krishnan bayley": "Krishna Bayley", "kishna bayley": "Krishna Bayley",
    "michael macpherson": "Michael MacPherson", "michelle": "Michelle (locum)",
    "ben": "Ben (locum)", "fiona forrester": "Fiona Forrester",
    "lauren mclean": "Lauren McLean", "laurn mclean": "Lauren McLean",
    "cynthia paredes": "Cynthia Paredes", "nikki pennifold": "Nikki Pennifold",
    "li chei lee": "Li Chei Lee",
}

ROWS = [
    (4, "ED_CB"), (5, "GPA_CB"), (6, "GPO_CB"), (7, "OBS_CB"),
    (9, "ED_JT"), (10, "GPA_JT"), (11, "GPO_JT"), (12, "OBS_JT"),
    (14, "ED_LAURA"), (16, "ED_ORROROO"),
]
MONTHS = [("2026-07", 3, 33), ("2026-08", 35, 65)]  # label, first col, last col


def canon(raw):
    s = " ".join(str(raw).split())
    if not s:
        return None
    if s.upper().startswith("DIVERT"):
        return "DIVERT"
    backup = None
    if s.endswith(")") and "(" in s:
        base, tag = s.rsplit("(", 1)
        tag = tag.rstrip(")").strip()
        if tag.upper() in ("SJ", "HD"):
            s, backup = base.strip(), tag.upper()
    key = s.lower()
    if key not in CANON:
        sys.exit(f"unrecognised name in roster: {raw!r}")
    return CANON[key] + (f"|{backup}" if backup else "")


def doc_id(name):
    return name.lower().replace(" ", "_").replace("'", "").replace("(", "").replace(")", "")


def main():
    src = sys.argv[1]
    wb = load_workbook(src, data_only=True)
    ws = wb["Sheet1"]

    out = []
    for label, c0, c1 in MONTHS:
        days = []
        for col in range(c0, c1 + 1):
            letter = get_column_letter(col)
            dow = ws[f"{letter}2"].value or ""
            day = col - c0 + 1
            row = {"date": f"{label}-{day:02d}", "weekday": dow[:3]}
            for r, role in ROWS:
                v = ws.cell(row=r, column=col).value
                name = canon(v) if v is not None else None
                if name and "|" in name:
                    name, bk = name.split("|")
                    row.setdefault("_backup", {})[role] = bk
                row[role] = doc_id(name) if name and name != "DIVERT" else (name or None)
            days.append(row)

        # Laura's line simply stops being rostered in August - drop the role
        # for that month rather than reporting 31 fake gaps
        roles = [role for _, role in ROWS]
        if all(d["ED_LAURA"] is None for d in days):
            roles = [r for r in roles if r != "ED_LAURA"]
            for d in days:
                del d["ED_LAURA"]

        names = {}
        for d in days:
            for r in roles:
                w = d.get(r)
                if w and w != "DIVERT":
                    names.setdefault(w, None)

        stats = {}
        for did in names:
            ed = mat = wk = multi = triple = 0
            streak = best = 0
            for d in days:
                todays = [r for r in roles if d.get(r) == did]
                if todays:
                    if any(r.startswith("ED") for r in todays):
                        ed += 1
                    if not all(r.startswith("ED") for r in todays):
                        mat += 1
                    if len(todays) > 1:
                        multi += 1
                    if len(todays) > 2:
                        triple += 1
                    if d["weekday"] in ("Sat", "Sun"):
                        wk += 1
                    streak += 1
                    best = max(best, streak)
                else:
                    streak = 0
            total = sum(1 for d in days if any(d.get(r) == did for r in roles))
            pretty = next(v for k, v in CANON.items() if doc_id(v) == did)
            stats[did] = {
                "name": pretty, "total_oncall_days": total, "ed_days": ed,
                "maternity_days": mat, "weekend_days": wk,
                "double_hatted_days": multi, "triple_hatted_days": triple,
                "longest_stretch": best, "leave_days": 0, "reduced_load": False,
            }

        gaps = sum(1 for d in days for r in roles if d.get(r) is None)
        ed_roles = [r for r in roles if r.startswith("ED")]
        ed_gaps = sum(1 for d in days for r in ed_roles if d.get(r) is None)
        divert = sum(1 for d in days for r in roles if d.get(r) == "DIVERT")
        multi_days = sum(s["double_hatted_days"] for s in stats.values())
        ed_slots = len(ed_roles) * len(days)

        out.append({
            "scenario": f"actual_{label[-2:]}",
            "label": f"ACTUAL roster — {label}",
            "model": "actual",
            "month": label, "active_site": None, "days": len(days),
            "roles": roles, "roster": days, "doctor_stats": stats,
            "summary": {
                "ed_slots": ed_slots, "ed_gaps": ed_gaps,
                "ed_fill_pct": round(100 * (ed_slots - ed_gaps) / ed_slots, 1),
                "total_gaps": gaps, "divert_days": divert,
                "clare_backup_days": 0, "double_hatted_days": multi_days,
            },
        })

    (HERE / "actuals.json").write_text(json.dumps(out, indent=1))
    for r in out:
        s = r["summary"]
        print(f"\n=== {r['label']} ({len(r['roles'])} lines/day) ===")
        print(f"  unfilled cells: {s['total_gaps']} (ED: {s['ed_gaps']})"
              f" | DIVERT days: {s['divert_days']}"
              f" | multi-hatted person-days: {s['double_hatted_days']}")
        for did, st in sorted(r["doctor_stats"].items(),
                              key=lambda kv: -kv[1]["total_oncall_days"]):
            print(f"  {st['name']:<20} {st['total_oncall_days']:>2}d "
                  f"(ED {st['ed_days']}, mat {st['maternity_days']}, "
                  f"multi {st['double_hatted_days']}, triple {st['triple_hatted_days']}, "
                  f"stretch {st['longest_stretch']}, wkend {st['weekend_days']})")


if __name__ == "__main__":
    main()
